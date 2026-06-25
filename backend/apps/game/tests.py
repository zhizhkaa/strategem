"""Regression tests for game API, formulas, models, and frontend assets."""

import json
import importlib.util
import tempfile
from io import BytesIO
from pathlib import Path

from apps.management.models import Faculty, Group, Team
from apps.game.engine import FormulaValidator, get_calculator, get_state_manager
from apps.game.engine.state_manager import DecisionState
from apps.game.models import Document, Game, GameDifficulty, GameStatus
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.test.utils import override_settings
from rest_framework import status
from rest_framework.test import APITestCase


class GameApiTests(APITestCase):
    """Covers the public game API and period decision workflow."""

    def setUp(self):
        self.faculty = Faculty.objects.create(name="ИТ")
        self.group = Group.objects.create(name="ИМК-101", faculty=self.faculty, year=2026)
        self.team_counter = 0
        self._set_admin_session()

    def _set_admin_session(self) -> None:
        session = self.client.session
        session["is_admin"] = True
        session.save()

    def _create_team(self) -> Team:
        self.team_counter += 1
        return Team.objects.create(name=f"Команда {self.team_counter}", group=self.group)

    def _create_game(
        self,
        difficulty: str = GameDifficulty.STANDARD,
        total_periods: int = 12,
    ) -> Game:
        team = self._create_team()
        response = self.client.post(
            "/api/games/",
            {
                "team": team.id,
                "difficulty": difficulty,
                "total_periods": total_periods,
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        return Game.objects.get(team=team)

    def test_create_game_uses_standard_profile_and_12_periods(self):
        team = self._create_team()

        response = self.client.post(
            "/api/games/",
            {
                "team": team.id,
                "difficulty": GameDifficulty.STANDARD,
                "total_periods": 12,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        game = Game.objects.get(team=team)
        period = game.periods.get(period_number=1)

        self.assertEqual(game.status, GameStatus.ACTIVE)
        self.assertEqual(game.difficulty, GameDifficulty.STANDARD)
        self.assertEqual(game.total_periods, 12)
        self.assertEqual(period.E7, 14500)
        self.assertEqual(period.E17, 12000)
        self.assertEqual(period.E19, 1500)
        self.assertEqual(period.TF1, 1000)
        self.assertEqual(period.TF5, 0)

    def test_create_game_with_higheq_profile_sets_initial_period(self):
        team = self._create_team()

        response = self.client.post(
            "/api/games/",
            {
                "team": team.id,
                "difficulty": GameDifficulty.HIGHEQ,
                "total_periods": 10,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        game = Game.objects.get(team=team)
        period = game.periods.get(period_number=1)

        self.assertEqual(period.P1, 400)
        self.assertEqual(period.P3, 38530)
        self.assertEqual(period.E17, 25000)
        self.assertEqual(period.F7, 0.8)
        self.assertEqual(period.TF1, 5000)

    def test_create_game_with_simple_profile_matches_pdf_initial_values(self):
        team = self._create_team()

        response = self.client.post(
            "/api/games/",
            {
                "team": team.id,
                "difficulty": GameDifficulty.SIMPLE,
                "total_periods": 10,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

        game = Game.objects.get(team=team)
        period = game.periods.get(period_number=1)

        self.assertEqual(period.E7, 15000)
        self.assertEqual(period.E17, 14500)
        self.assertEqual(period.E19, 500)
        self.assertEqual(period.TF1, 0)
        self.assertEqual(period.TF4, 300)
        self.assertEqual(period.TF5, 1000)

    def test_list_and_detail_include_difficulty(self):
        team = self._create_team()
        game = Game.objects.create(
            team=team,
            status=GameStatus.ACTIVE,
            difficulty=GameDifficulty.TOUGH,
            total_periods=10,
        )

        list_response = self.client.get("/api/games/")
        detail_response = self.client.get(f"/api/games/{game.id}/")

        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)
        self.assertEqual(list_response.data[0]["difficulty"], GameDifficulty.TOUGH)
        self.assertEqual(list_response.data[0]["difficulty_display"], "Tough")
        self.assertEqual(detail_response.data["difficulty"], GameDifficulty.TOUGH)
        self.assertEqual(detail_response.data["difficulty_display"], "Tough")

    def test_archived_games_are_hidden_from_default_list(self):
        active_game = self._create_game()
        archived_game = self._create_game()

        response = self.client.post(f"/api/games/{archived_game.id}/archive/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        archived_game.refresh_from_db()
        self.assertTrue(archived_game.is_archived)
        self.assertIsNotNone(archived_game.archived_at)

        default_response = self.client.get("/api/games/")
        archive_response = self.client.get("/api/games/?archived=1")
        all_response = self.client.get("/api/games/?archived=all")

        self.assertEqual(
            [game_data["id"] for game_data in default_response.data],
            [active_game.id],
        )
        self.assertEqual(
            [game_data["id"] for game_data in archive_response.data],
            [archived_game.id],
        )
        self.assertCountEqual(
            [game_data["id"] for game_data in all_response.data],
            [active_game.id, archived_game.id],
        )

        restore_response = self.client.post(f"/api/games/{archived_game.id}/unarchive/")
        self.assertEqual(restore_response.status_code, status.HTTP_200_OK)
        archived_game.refresh_from_db()
        self.assertFalse(archived_game.is_archived)
        self.assertIsNone(archived_game.archived_at)

    def test_game_excel_export_returns_workbook(self):
        from openpyxl import load_workbook

        game = self._create_game()

        response = self.client.get(f"/api/games/{game.id}/export/excel/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(response.content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Сводные итоги", "Периоды"])
        self.assertEqual(workbook["Периоды"]["A11"].value, "Код")
        self.assertEqual(workbook["Периоды"]["C11"].value, "Период 1")
        self.assertEqual(workbook["Сводные итоги"]["A11"].value, "Показатель")
        self.assertEqual(workbook["Сводные итоги"]["B11"].value, "Период 1")
        self.assertEqual(workbook["Сводные итоги"]["A12"].value, "Сводный счёт")

    def test_game_pdf_export_returns_pdf_when_reportlab_is_installed(self):
        if importlib.util.find_spec("reportlab") is None:
            self.skipTest("reportlab is not installed in the current environment")

        game = self._create_game()

        response = self.client.get(f"/api/games/{game.id}/export/pdf/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_rejects_unsupported_15_periods(self):
        team = self._create_team()

        response = self.client.post(
            "/api/games/",
            {
                "team": team.id,
                "difficulty": GameDifficulty.STANDARD,
                "total_periods": 15,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("total_periods", response.data)

    def test_team_password_is_required_to_enter_game(self):
        team = Team.objects.create(
            name="Команда с паролем",
            group=self.group,
            access_password="luna-42",
        )
        Game.objects.create(
            team=team,
            status=GameStatus.ACTIVE,
            difficulty=GameDifficulty.STANDARD,
            total_periods=12,
        )

        session = self.client.session
        session["is_admin"] = False
        session.save()

        missing_response = self.client.post(f"/api/teams/{team.id}/game/", {}, format="json")
        wrong_response = self.client.post(
            f"/api/teams/{team.id}/game/",
            {"password": "wrong"},
            format="json",
        )
        ok_response = self.client.post(
            f"/api/teams/{team.id}/game/",
            {"password": "luna-42"},
            format="json",
        )

        self.assertEqual(missing_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(wrong_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(ok_response.status_code, status.HTTP_200_OK)
        self.assertEqual(ok_response.data["game"]["team"], team.id)

    def test_team_without_password_can_enter_game_without_password(self):
        team = Team.objects.create(
            name="Команда без пароля",
            group=self.group,
            access_password="",
        )
        Game.objects.create(
            team=team,
            status=GameStatus.ACTIVE,
            difficulty=GameDifficulty.STANDARD,
            total_periods=12,
        )

        session = self.client.session
        session["is_admin"] = False
        session.save()

        response = self.client.post(f"/api/teams/{team.id}/game/", {}, format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["game"]["team"], team.id)

    def test_team_can_enter_finished_game(self):
        team = Team.objects.create(
            name="Команда с завершённой игрой",
            group=self.group,
            access_password="terra-24",
        )
        Game.objects.create(
            team=team,
            status=GameStatus.FINISHED,
            difficulty=GameDifficulty.STANDARD,
            total_periods=12,
        )

        session = self.client.session
        session["is_admin"] = False
        session.save()

        response = self.client.post(
            f"/api/teams/{team.id}/game/",
            {"password": "terra-24"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["game"]["team"], team.id)
        self.assertEqual(response.data["game"]["status"], GameStatus.FINISHED)

    def test_team_list_never_returns_raw_password(self):
        team = Team.objects.create(
            name="Команда пароль",
            group=self.group,
            access_password="atlas-17",
        )

        session = self.client.session
        session["is_admin"] = False
        session.save()
        public_response = self.client.get("/api/teams/")

        session = self.client.session
        session["is_admin"] = True
        session.save()
        admin_response = self.client.get("/api/teams/")

        self.assertEqual(public_response.status_code, status.HTTP_200_OK)
        self.assertNotIn("access_password", public_response.data[0])
        self.assertTrue(public_response.data[0]["has_access_password"])
        self.assertEqual(admin_response.status_code, status.HTTP_200_OK)
        self.assertEqual(admin_response.data[0]["id"], team.id)
        self.assertNotIn("access_password", admin_response.data[0])
        self.assertTrue(admin_response.data[0]["has_access_password"])

    def test_team_list_reports_finished_game_status(self):
        active_team = Team.objects.create(
            name="Активная команда",
            group=self.group,
            access_password="atlas-17",
        )
        finished_team = Team.objects.create(
            name="Завершившая команда",
            group=self.group,
            access_password="terra-24",
        )
        Game.objects.create(
            team=active_team,
            status=GameStatus.ACTIVE,
            difficulty=GameDifficulty.STANDARD,
            total_periods=12,
        )
        Game.objects.create(
            team=finished_team,
            status=GameStatus.FINISHED,
            difficulty=GameDifficulty.STANDARD,
            total_periods=12,
        )

        response = self.client.get("/api/teams/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        by_id = {team_data["id"]: team_data for team_data in response.data}
        self.assertTrue(by_id[active_team.id]["has_active_game"])
        self.assertEqual(by_id[active_team.id]["game_status"], GameStatus.ACTIVE)
        self.assertFalse(by_id[finished_team.id]["has_active_game"])
        self.assertTrue(by_id[finished_team.id]["has_finished_game"])
        self.assertEqual(by_id[finished_team.id]["game_status"], GameStatus.FINISHED)

    def test_api_docs_require_admin_session(self):
        session = self.client.session
        session["is_admin"] = False
        session.save()

        docs_response = self.client.get("/api/docs/")
        schema_response = self.client.get("/api/schema/")

        self.assertEqual(docs_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(schema_response.status_code, status.HTTP_403_FORBIDDEN)

        self._set_admin_session()
        admin_docs_response = self.client.get("/api/docs/")

        self.assertEqual(admin_docs_response.status_code, status.HTTP_200_OK)

    def test_admin_can_update_team_password(self):
        team = Team.objects.create(
            name="Команда пароль update",
            group=self.group,
            access_password="atlas-17",
        )

        response = self.client.patch(
            f"/api/teams/{team.id}/",
            {"access_password": "nova-55"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        team.refresh_from_db()
        self.assertNotEqual(team.access_password, "nova-55")
        self.assertTrue(team.check_access_password("nova-55"))
        self.assertNotIn("access_password", response.data)
        self.assertTrue(response.data["has_access_password"])

    def test_admin_can_disable_team_password(self):
        team = Team.objects.create(
            name="Команда отключить пароль",
            group=self.group,
            access_password="atlas-17",
        )

        response = self.client.patch(
            f"/api/teams/{team.id}/",
            {"password_enabled": False, "access_password": ""},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        team.refresh_from_db()
        self.assertEqual(team.access_password, "")
        self.assertFalse(response.data["has_access_password"])

    def test_admin_can_keep_existing_team_password_with_empty_patch_value(self):
        team = Team.objects.create(
            name="Команда оставить пароль",
            group=self.group,
            access_password="atlas-17",
        )
        stored_password = team.access_password

        response = self.client.patch(
            f"/api/teams/{team.id}/",
            {"password_enabled": True, "access_password": ""},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        team.refresh_from_db()
        self.assertEqual(team.access_password, stored_password)
        self.assertTrue(team.check_access_password("atlas-17"))
        self.assertTrue(response.data["has_access_password"])

    def test_admin_cannot_save_too_short_team_password(self):
        team = Team.objects.create(
            name="Команда короткий пароль",
            group=self.group,
            access_password="atlas-17",
        )

        response = self.client.patch(
            f"/api/teams/{team.id}/",
            {"access_password": "a"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("access_password", response.data)

    def test_decision_structure_treats_residual_decisions_as_manual_inputs(self):
        response = self.client.get("/api/games/decision-structure/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        first_group = response.data["summary_groups"][0]
        population_decision = response.data["ministers"]["population"]["decisions"][0]

        self.assertIn("P9", first_group["inputs"])
        self.assertIn("P10", first_group["inputs"])
        self.assertNotIn("P10", first_group["auto"])
        self.assertIn("P10", population_decision["inputs"])
        self.assertNotIn("P10", population_decision["auto"])

    def test_batch_accepts_manual_residual_decision_values(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 1900, "P10": 1400}, "minister": "population"},
            format="json",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )
        self.assertTrue(response.data["success"])

        period = game.periods.get(period_number=1)
        self.assertEqual(period.P9, 1900)
        self.assertEqual(period.P10, 1400)
        self.assertIn("P10", period.user_inputs)

    def test_goods_distribution_is_available_without_food_distribution(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        response = self.client.get(f"/api/games/{game.id}/state/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        stages = {stage["key"]: stage for stage in response.data["available_stages"]}
        self.assertTrue(stages["food_distribution"]["available"])
        self.assertTrue(stages["goods_distribution"]["available"])
        self.assertEqual(response.data["parameters"]["P11"]["status"], "next_to_fill")
        self.assertEqual(response.data["parameters"]["P12"]["status"], "next_to_fill")
        self.assertEqual(response.data["parameters"]["P13"]["status"], "next_to_fill")

    def test_batch_reports_food_distribution_underallocation(self):
        game = self._create_game(
            difficulty=GameDifficulty.STANDARD,
            total_periods=10,
        )

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 3000, "P10": 200}, "minister": "population"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["success"])
        self.assertIn("P9", response.data["errors"])
        self.assertIn("P10", response.data["errors"])
        self.assertEqual(response.data["errors"]["P9"], "Баланс P2 не соблюдается")
        self.assertEqual(response.data["errors"]["P10"], "Баланс P2 не соблюдается")
        self.assertEqual(response.data["message"], "Баланс P2 не соблюдается")
        self.assertNotIn("E20", response.data["errors"])
        self.assertNotIn("F15", response.data["errors"])
        self.assertIn("P9", response.data["validation_state"]["errors"])
        self.assertIn("P10", response.data["validation_state"]["errors"])
        self.assertNotIn("E20", response.data["validation_state"]["errors"])
        self.assertNotIn("F15", response.data["validation_state"]["errors"])

        game.refresh_from_db()
        self.assertEqual(game.decision_finance, 0)
        self.assertEqual(game.decision_capital, 0)

    def test_batch_accepts_energy_production_matching_displayed_e23(self):
        game = self._create_game(
            difficulty=GameDifficulty.STANDARD,
            total_periods=10,
        )
        game.decision_finance = 3
        game.decision_energy = 1
        game.save()

        period = game.get_current_period_obj()
        period.E7 = 12206.999
        period.E20 = 0
        period.E21 = 0
        period.E22 = 0
        period.E23 = 12206.999
        period.user_inputs = ["E20", "E21", "E22", "E23"]
        period.save()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"E24": 12000, "E25": 207}, "minister": "energy"},
            format="json",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )
        self.assertTrue(response.data["success"], msg=response.data)
        self.assertEqual(response.data["errors"], {})

    def test_first_period_industry_uses_initial_capital_requirements(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        state_response = self.client.get(f"/api/games/{game.id}/state/")
        self.assertEqual(state_response.status_code, status.HTTP_200_OK)

        stages = {
            stage["key"]: stage for stage in state_response.data["available_stages"]
        }
        self.assertTrue(stages["industry_investment"]["available"])
        self.assertEqual(
            state_response.data["parameters"]["G18"]["status"],
            "next_to_fill",
        )

    def test_first_period_trade_finance_uses_initial_requirements(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        state_response = self.client.get(f"/api/games/{game.id}/state/")
        self.assertEqual(state_response.status_code, status.HTTP_200_OK)

        stages = {
            stage["key"]: stage for stage in state_response.data["available_stages"]
        }
        self.assertTrue(stages["debt_payment"]["available"])
        self.assertTrue(stages["import_distribution"]["available"])
        self.assertEqual(
            state_response.data["parameters"]["TF14"]["status"],
            "next_to_fill",
        )
        self.assertEqual(
            state_response.data["parameters"]["TF16"]["status"],
            "next_to_fill",
        )

    def test_state_marks_dynamic_zero_bound_decisions_as_not_fixed(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        state_response = self.client.get(f"/api/games/{game.id}/state/")
        self.assertEqual(state_response.status_code, status.HTTP_200_OK)

        parameters = state_response.data["parameters"]
        self.assertTrue(parameters["E20"]["is_fixed"])
        for param in (
            "E27",
            "G18",
            "G19",
            "F14",
            "F15",
            "TF14",
            "TF15",
            "TF16",
            "TF17",
            "TF18",
        ):
            with self.subTest(param=param):
                self.assertFalse(parameters[param]["is_fixed"])

    def test_first_period_energy_can_submit_e27_before_industry_and_agriculture(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        population_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {
                "parameters": {
                    "P9": 1900,
                    "P10": 1400,
                    "P11": 2000,
                    "P12": 1500,
                    "P13": 0,
                },
                "minister": "population",
            },
            format="json",
        )
        self.assertEqual(
            population_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(population_response, "data", population_response.content),
        )
        self.assertTrue(
            population_response.data["success"],
            msg=population_response.data,
        )

        energy_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"E26": 500, "E27": 300}, "minister": "energy"},
            format="json",
        )

        self.assertEqual(
            energy_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(energy_response, "data", energy_response.content),
        )
        self.assertTrue(energy_response.data["success"], msg=energy_response.data)

        period = game.get_current_period_obj()
        period.refresh_from_db()
        self.assertEqual(period.E26, 500)
        self.assertEqual(period.E27, 300)

    def test_first_period_industry_can_submit_before_energy_investment(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        population_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {
                "parameters": {
                    "P9": 1900,
                    "P10": 1400,
                    "P11": 2000,
                    "P12": 1500,
                    "P13": 0,
                },
                "minister": "population",
            },
            format="json",
        )
        self.assertEqual(
            population_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(population_response, "data", population_response.content),
        )
        self.assertTrue(
            population_response.data["success"],
            msg=population_response.data,
        )

        industry_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"G18": 100, "G19": 200}, "minister": "industry"},
            format="json",
        )

        self.assertEqual(
            industry_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(industry_response, "data", industry_response.content),
        )
        self.assertTrue(
            industry_response.data["success"],
            msg=industry_response.data,
        )

        game.refresh_from_db()
        period = game.get_current_period_obj()
        self.assertEqual(period.G18, 100)
        self.assertEqual(period.G19, 200)
        self.assertEqual(game.decision_capital, 3)

        state_response = self.client.get(f"/api/games/{game.id}/state/")
        stages = {
            stage["key"]: stage for stage in state_response.data["available_stages"]
        }
        self.assertTrue(stages["industry_investment"]["completed"])
        self.assertFalse(stages["energy_investment"]["completed"])

    def test_batch_recalculates_hidden_fixed_currency_fields_before_validation(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        game.decision_capital = 1
        game.decision_energy = 1
        game.decision_finance = 3
        game.decision_import = 3
        game.save()

        period = game.get_current_period_obj()
        period.TF10 = 0
        period.TF11 = 0
        period.TF12 = 0
        period.E7 = 14500
        period.user_inputs = ["TF10", "TF11", "TF12"]
        period.save()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {
                "parameters": {
                    "P9": 3000,
                    "P10": 300,
                    "P11": 1000,
                    "P12": 1000,
                    "P13": 1500,
                    "E21": 1000,
                    "E22": 1000,
                    "E23": 12300,
                    "E24": 12000,
                    "E25": 300,
                    "E26": 0,
                    "E27": 0,
                    "G18": 0,
                    "G19": 0,
                    "F14": 500,
                    "F15": 500,
                    "TF13": 0,
                    "TF14": 0,
                    "TF15": 0,
                    "TF16": 0,
                    "TF17": 0,
                    "TF18": 0,
                }
            },
            format="json",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )
        self.assertTrue(response.data["success"], msg=response.data)
        self.assertNotIn("TF10", response.data["errors"])
        self.assertEqual(response.data["auto_calculated"]["TF10"], 1000.0)
        self.assertEqual(response.data["auto_calculated"]["TF11"], 1500.0)
        self.assertEqual(response.data["auto_calculated"]["TF12"], 300.0)

        period.refresh_from_db()
        self.assertEqual(period.TF10, 1000)
        self.assertEqual(period.TF11, 1500)
        self.assertEqual(period.TF12, 300)
        self.assertNotIn("TF10", period.user_inputs)

    def test_batch_parameter_submission_returns_blocking_incomplete_state(self):
        game = self._create_game()
        period = game.get_current_period_obj()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 1000}, "minister": "population"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["success"])
        self.assertEqual(response.data["updated_parameters"]["P9"], 1000)
        self.assertIn("validation_state", response.data)
        self.assertEqual(response.data["validation_state"]["errors"], [])

        expected_incomplete = {
            "P10",
            "P11",
            "P12",
            "P13",
            "E20",
            "E21",
            "E22",
            "E23",
            "E24",
            "E25",
            "E26",
            "E27",
            "G18",
            "G19",
            "F14",
            "F15",
            "TF13",
            "TF14",
            "TF15",
            "TF16",
            "TF17",
            "TF18",
        }
        actual_incomplete = set(response.data["validation_state"]["incomplete"])

        self.assertSetEqual(actual_incomplete, expected_incomplete)
        self.assertIn("population", response.data["validation_state"]["by_minister"])
        self.assertIn(
            "P10",
            response.data["validation_state"]["by_minister"]["population"]["incomplete"],
        )

        period.refresh_from_db()
        self.assertSetEqual(
            set(period.validation_state["incomplete"]),
            expected_incomplete,
        )

    def test_calculator_reset_restores_initial_period_when_periods_were_cleared(self):
        game = self._create_game(total_periods=10)
        game.periods.all().delete()

        reset_response = self.client.post(f"/api/admin/calculator/{game.id}/reset/")
        state_response = self.client.get(f"/api/admin/calculator/{game.id}/state/")

        self.assertEqual(reset_response.status_code, status.HTTP_200_OK)
        self.assertEqual(state_response.status_code, status.HTTP_200_OK)
        self.assertEqual(state_response.data["current_period"], 1)
        self.assertEqual(len(state_response.data["periods"]), 1)
        self.assertEqual(state_response.data["periods"][0]["period_number"], 1)

        game.refresh_from_db()
        self.assertEqual(game.current_period, 1)
        self.assertEqual(game.periods.count(), 1)

    def test_admin_calculator_preserves_completed_decision_snapshot(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period_1_inputs = {
            "P9": 1900,
            "P10": 1400,
            "P11": 2000,
            "P12": 1500,
            "P13": 0,
            "E20": 400,
            "E21": 0,
            "E22": 0,
            "E23": 14600,
            "E24": 8000,
            "E25": 6600,
            "E26": 500,
            "E27": 300,
            "G18": 100,
            "G19": 400,
            "F14": 200,
            "F15": 0,
            "TF9": 0,
            "TF10": 0,
            "TF11": 0,
            "TF12": 1400,
            "TF13": 1000,
            "TF14": 0,
            "TF15": 2400,
            "TF16": 1000,
            "TF17": 1400,
            "TF18": 0,
        }

        response = self.client.post(
            f"/api/admin/calculator/{game.id}/calculate/",
            {"parameters": period_1_inputs},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        game.refresh_from_db()
        completed_period = game.periods.get(period_number=1)
        next_period = game.periods.get(period_number=2)

        self.assertEqual(game.current_period, 2)
        self.assertEqual(completed_period.P10, 1400)
        self.assertEqual(completed_period.P13, 0)
        self.assertEqual(completed_period.TF12, 1400)
        self.assertEqual(completed_period.TF15, 2400)
        self.assertEqual(next_period.P2, 2811)
        self.assertEqual(next_period.P9, 0)
        self.assertNotEqual(next_period.P10, 911)
        self.assertEqual(next_period.P10, 0)
        self.assertEqual(next_period.P11, 0)

    def test_admin_calculator_recalculates_fixed_currency_fields_before_validation(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period_1_inputs = {
            "P9": 1900,
            "P10": 1400,
            "P11": 2000,
            "P12": 1500,
            "P13": 0,
            "E20": 400,
            "E21": 0,
            "E22": 0,
            "E23": 14600,
            "E24": 8000,
            "E25": 6600,
            "E26": 500,
            "E27": 300,
            "G18": 100,
            "G19": 400,
            "F14": 200,
            "F15": 0,
            "TF13": 1000,
            "TF14": 0,
            "TF15": 2400,
            "TF16": 1000,
            "TF17": 1400,
            "TF18": 0,
        }

        response = self.client.post(
            f"/api/admin/calculator/{game.id}/calculate/",
            {"parameters": period_1_inputs},
            format="json",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )

        completed_period = game.periods.get(period_number=1)
        self.assertEqual(completed_period.TF10, 0)
        self.assertEqual(completed_period.TF11, 0)
        self.assertEqual(completed_period.TF12, 1400)

    def test_admin_calculator_rejects_capital_investment_overrun(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        overspent_inputs = {
            "P9": 0,
            "P10": 3300,
            "P11": 0,
            "P12": 1000,
            "P13": 2500,
            "E20": 0,
            "E21": 0,
            "E22": 0,
            "E23": 15000,
            "E24": 0,
            "E25": 15000,
            "E26": 700,
            "E27": 400,
            "G18": 200,
            "G19": 100,
            "F14": 115,
            "F15": 0,
            "TF13": 0,
            "TF14": 0,
            "TF15": 0,
            "TF16": 0,
            "TF17": 0,
            "TF18": 0,
        }

        response = self.client.post(
            f"/api/admin/calculator/{game.id}/calculate/",
            {"parameters": overspent_inputs},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("F14", response.data["errors"])

        period = game.periods.get(period_number=1)
        self.assertEqual(period.F15, 0)

    def test_next_period_preserves_completed_decisions_and_resets_new_decisions(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period_1_inputs = [
            ("P9", 1900),
            ("P10", 1400),
            ("TF12", 1400),
            ("P11", 2000),
            ("P12", 1500),
            ("P13", 0),
            ("E20", 400),
            ("E21", 0),
            ("E22", 0),
            ("E23", 14600),
            ("E24", 8000),
            ("E25", 6600),
            ("E26", 500),
            ("G18", 100),
            ("F14", 200),
            ("E27", 300),
            ("G19", 400),
            ("F15", 0),
            ("TF13", 1000),
            ("TF14", 0),
            ("TF15", 2400),
            ("TF16", 1000),
            ("TF17", 1400),
            ("TF18", 0),
        ]

        for param, value in period_1_inputs:
            response = self.client.post(
                f"/api/games/{game.id}/parameters/{param}/",
                {"value": value, "force": True},
                format="json",
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_200_OK,
                f"{param}: {response.data}",
            )

        period_1 = game.periods.get(period_number=1)
        self.assertEqual(period_1.P10, 1400)
        self.assertEqual(period_1.TF12, 1400)

        response = self.client.post(f"/api/games/{game.id}/next-period/", format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        game.refresh_from_db()
        completed_period = game.periods.get(period_number=1)
        new_period = game.periods.get(period_number=2)

        self.assertEqual(game.current_period, 2)
        self.assertEqual(completed_period.P9, 1900)
        self.assertEqual(completed_period.P10, 1400)
        self.assertEqual(completed_period.TF12, 1400)
        self.assertEqual(new_period.P9, 0)
        self.assertEqual(new_period.P10, 0)
        self.assertEqual(new_period.P11, 0)
        self.assertEqual(new_period.P12, 0)
        self.assertEqual(new_period.TF12, 0)
        self.assertNotEqual(new_period.P10, 911)

        state_response = self.client.get(f"/api/games/{game.id}/state/")
        self.assertEqual(state_response.status_code, status.HTTP_200_OK)
        stages = {stage["key"]: stage for stage in state_response.data["available_stages"]}
        for stage_key in (
            "food_distribution",
            "goods_distribution",
            "energy_distribution",
            "energy_production",
            "energy_investment",
            "industry_investment",
            "agriculture_investment",
            "loan",
            "debt_payment",
            "import_distribution",
        ):
            with self.subTest(stage=stage_key):
                self.assertTrue(stages[stage_key]["available"])
        self.assertEqual(state_response.data["parameters"]["P9"]["status"], "next_to_fill")
        self.assertEqual(state_response.data["parameters"]["P10"]["status"], "next_to_fill")
        self.assertEqual(state_response.data["parameters"]["P11"]["status"], "next_to_fill")
        self.assertEqual(state_response.data["parameters"]["E21"]["status"], "next_to_fill")
        self.assertEqual(state_response.data["parameters"]["TF13"]["status"], "next_to_fill")
        self.assertEqual(state_response.data["validation_state"]["errors"], [])
        self.assertEqual(state_response.data["validation_state"]["incomplete"], [])

    def test_batch_ignores_unchanged_already_filled_parameters(self):
        game = self._create_game(
            difficulty=GameDifficulty.STANDARD,
            total_periods=10,
        )

        first_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 3300, "P10": 0}, "minister": "population"},
            format="json",
        )
        self.assertEqual(first_response.status_code, status.HTTP_200_OK)

        second_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 3300, "P11": 2000, "P12": 1500}, "minister": "population"},
            format="json",
        )

        self.assertEqual(
            second_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(second_response, "data", second_response.content),
        )
        period = game.periods.get(period_number=1)
        self.assertEqual(period.P9, 3300)
        self.assertEqual(period.P11, 2000)
        self.assertEqual(period.P12, 1500)

    def test_batch_allows_editing_already_filled_parameters(self):
        game = self._create_game(
            difficulty=GameDifficulty.STANDARD,
            total_periods=10,
        )

        first_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 3300, "P10": 0}, "minister": "population"},
            format="json",
        )
        self.assertEqual(first_response.status_code, status.HTTP_200_OK)

        second_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P11": 2000, "P12": 1500}, "minister": "population"},
            format="json",
        )
        self.assertEqual(second_response.status_code, status.HTTP_200_OK)

        edit_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 3200}, "minister": "population"},
            format="json",
        )

        self.assertEqual(
            edit_response.status_code,
            status.HTTP_200_OK,
            msg=getattr(edit_response, "data", edit_response.content),
        )
        period = game.periods.get(period_number=1)
        self.assertEqual(period.P9, 3200)
        self.assertEqual(period.P10, 0)

    def test_next_period_repairs_existing_negative_f15_before_using_history(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period = game.periods.get(period_number=1)
        period.F6 = 900
        period.F4 = 150
        period.F15 = -515
        period.save()

        game.advance_period()

        period.refresh_from_db()
        next_period = game.periods.get(period_number=2)
        self.assertEqual(period.F15, 0)
        self.assertEqual(next_period.F6, 750)

    def test_validate_period_reports_capital_investment_overrun(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period = game.periods.get(period_number=1)
        period.P12 = 1000
        period.E26 = 700
        period.E27 = 400
        period.G18 = 200
        period.G19 = 100
        period.F14 = 115
        period.F15 = 0
        period.user_inputs = ["E26", "E27", "G18", "G19", "F14"]
        period.save()

        response = self.client.post(f"/api/games/{game.id}/validate-period/", format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["valid"])
        self.assertFalse(response.data["can_advance"])
        self.assertIn("F14", response.data["errors"])

    def test_validate_period_reports_all_decision_residual_overruns(self):
        cases = [
            (
                {"P2": 100, "P9": 101},
                ["P9"],
            ),
            (
                {"P3": 100, "P11": 60, "P12": 41},
                ["P11", "P12"],
            ),
            (
                {"E7": 100, "E20": 20, "E21": 40, "E22": 41},
                ["E20", "E21", "E22"],
            ),
            (
                {"E23": 100, "E24": 101},
                ["E24"],
            ),
            (
                {"TF9": 0, "TF10": 10, "TF11": 10, "TF12": 10, "TF13": 10, "TF14": 41},
                ["TF14"],
            ),
            (
                {"TF15": 100, "TF16": 60, "TF17": 41},
                ["TF16", "TF17"],
            ),
        ]

        for overrides, expected_errors in cases:
            with self.subTest(overrides=overrides):
                game = self._create_game(
                    difficulty=GameDifficulty.SIMPLE,
                    total_periods=10,
                )
                period = game.periods.get(period_number=1)
                for param_name, value in overrides.items():
                    setattr(period, param_name, value)
                period.user_inputs = list(overrides.keys())
                period.save()

                response = self.client.post(
                    f"/api/games/{game.id}/validate-period/",
                    format="json",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertFalse(response.data["valid"])
                for param_name in expected_errors:
                    self.assertIn(param_name, response.data["errors"])

    def test_next_period_rejects_capital_investment_overrun(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        game.decision_capital = 4
        game.decision_energy = 1
        game.decision_finance = 3
        game.decision_import = 3
        game.save()

        period = game.periods.get(period_number=1)
        period.P12 = 1000
        period.E26 = 700
        period.E27 = 400
        period.G18 = 200
        period.G19 = 100
        period.F14 = 115
        period.F15 = 0
        period.user_inputs = [
            "P9",
            "P11",
            "P12",
            "E21",
            "E22",
            "E24",
            "E26",
            "E27",
            "G18",
            "G19",
            "F14",
            "TF13",
            "TF14",
            "TF16",
            "TF17",
        ]
        period.save()

        response = self.client.post(f"/api/games/{game.id}/next-period/", format="json")

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(game.periods.count(), 1)
        self.assertIn("F14", response.data["validation_state"]["errors"])

    def test_force_parameter_submission_preserves_explicit_fixed_input(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )

        response = self.client.post(
            f"/api/games/{game.id}/parameters/P11/",
            {"value": 2300, "force": True},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        response = self.client.post(
            f"/api/games/{game.id}/parameters/E20/",
            {"value": 458, "force": True},
            format="json",
        )
        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )

        period = game.periods.get(period_number=1)
        self.assertEqual(period.E20, 458)
        self.assertNotEqual(period.E20, 460)

    def test_next_period_allows_forced_snapshot_values_that_differ_from_fixed_formulas(self):
        game = self._create_game(
            difficulty=GameDifficulty.SIMPLE,
            total_periods=10,
        )
        period_inputs = [
            ("P9", 1000),
            ("P10", 2300),
            ("P11", 2300),
            ("P12", 1000),
            ("P13", 200),
            ("E20", 458),
            ("E21", 0),
            ("E22", 0),
            ("E23", 14542),
            ("E24", 1000),
            ("E25", 13542),
            ("E26", 100),
            ("G18", 100),
            ("F14", 100),
            ("E27", 100),
            ("G19", 100),
            ("F15", 500),
            ("TF10", 0),
            ("TF11", 200),
            ("TF12", 2300),
            ("TF13", 0),
            ("TF14", 0),
            ("TF15", 2500),
            ("TF16", 1000),
            ("TF17", 1000),
            ("TF18", 500),
        ]

        for param, value in period_inputs:
            response = self.client.post(
                f"/api/games/{game.id}/parameters/{param}/",
                {"value": value, "force": True},
                format="json",
            )
            self.assertEqual(
                response.status_code,
                status.HTTP_200_OK,
                msg=f"{param}: {getattr(response, 'data', response.content)}",
            )

        period = game.periods.get(period_number=1)
        self.assertEqual(period.E20, 458)

        response = self.client.post(f"/api/games/{game.id}/next-period/", format="json")
        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "data", response.content),
        )

    def test_validate_period_returns_blocking_incomplete_by_minister(self):
        game = self._create_game()

        batch_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 1000}, "minister": "population"},
            format="json",
        )
        self.assertEqual(batch_response.status_code, status.HTTP_200_OK)

        response = self.client.post(f"/api/games/{game.id}/validate-period/", format="json")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["valid"])
        self.assertFalse(response.data["can_advance"])
        self.assertEqual(response.data["errors"], [])
        self.assertIn("validation_state", response.data)
        self.assertIn("population", response.data["validation_state"]["by_minister"])
        self.assertIn(
            "P10",
            response.data["validation_state"]["by_minister"]["population"]["incomplete"],
        )
        self.assertGreater(len(response.data["validation_state"]["incomplete"]), 0)

    def test_batch_validation_state_matches_validate_period_contract(self):
        game = self._create_game()

        batch_response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P9": 1000}, "minister": "population"},
            format="json",
        )
        self.assertEqual(batch_response.status_code, status.HTTP_200_OK)

        validate_response = self.client.post(
            f"/api/games/{game.id}/validate-period/",
            format="json",
        )
        self.assertEqual(validate_response.status_code, status.HTTP_200_OK)

        self.assertEqual(
            batch_response.data["validation_state"]["errors"],
            validate_response.data["validation_state"]["errors"],
        )
        self.assertEqual(
            batch_response.data["validation_state"]["incomplete"],
            validate_response.data["validation_state"]["incomplete"],
        )
        self.assertEqual(
            batch_response.data["validation_state"]["by_minister"],
            validate_response.data["validation_state"]["by_minister"],
        )

    def test_batch_rejects_operator_controlled_foreign_aid(self):
        game = self._create_game()
        period = game.get_current_period_obj()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"TF9": 100}, "minister": "trade_finance"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data["success"])
        self.assertIn("TF9", response.data["errors"])

        period.refresh_from_db()
        self.assertEqual(period.TF9, 0)

    def test_operator_controlled_param_requires_admin_even_with_force(self):
        game = self._create_game()
        session = self.client.session
        session["is_admin"] = False
        session.save()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/TF9/",
            {"value": 100, "force": True},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertIn("оператор", response.data["error"])

        period = game.get_current_period_obj()
        period.refresh_from_db()
        self.assertEqual(period.TF9, 0)

    def test_admin_can_set_operator_controlled_foreign_aid(self):
        game = self._create_game()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/TF9/",
            {"value": 100, "force": True},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        period = game.get_current_period_obj()
        period.refresh_from_db()
        self.assertEqual(period.TF9, 100)

    def test_decision_structure_exposes_foreign_aid_as_summary_info_and_trade_context(self):
        response = self.client.get("/api/games/decision-structure/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        summary_inputs = {
            param
            for group in response.data["summary_groups"]
            for param in group["inputs"]
        }
        trade_finance_inputs = {
            param
            for decision in response.data["ministers"]["trade_finance"]["decisions"]
            for param in decision["inputs"]
        }
        trade_finance_decision_keys = {
            decision["key"]
            for decision in response.data["ministers"]["trade_finance"]["decisions"]
        }
        group_numbers = {group["number"] for group in response.data["summary_groups"]}

        self.assertNotIn("TF9", summary_inputs)
        self.assertNotIn("TF9", trade_finance_inputs)
        self.assertNotIn("currency_revenue", trade_finance_decision_keys)
        self.assertNotIn(6, group_numbers)
        self.assertEqual(response.data["summary_info"], ["TF9"])
        self.assertIn("TF9", response.data["ministers"]["trade_finance"]["info_first"])
        self.assertIn("TF9", response.data["parameters"])
        self.assertNotIn("summary_situation", response.data)
        self.assertNotIn("TF10", response.data["summary_info"])
        self.assertNotIn("TF11", response.data["summary_info"])
        self.assertNotIn("TF12", response.data["summary_info"])

    def test_batch_is_atomic_when_dependent_value_is_invalid(self):
        game = self._create_game()
        period = game.get_current_period_obj()

        response = self.client.post(
            f"/api/games/{game.id}/parameters/batch/",
            {"parameters": {"P11": 3000, "P12": 600}, "minister": "population"},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(response.data["success"])
        self.assertIn("P12", response.data["errors"])

        period.refresh_from_db()
        self.assertEqual(period.P11, 0)
        self.assertEqual(period.P12, 0)


class GameModelFallbackTests(TestCase):
    """Checks model fallbacks that initialize periods without API helpers."""

    def setUp(self):
        faculty = Faculty.objects.create(name="Экономика")
        group = Group.objects.create(name="ЭК-201", faculty=faculty, year=2026)
        self.team = Team.objects.create(name="Команда fallback", group=group)

    def test_get_current_period_obj_uses_game_difficulty_profile(self):
        game = Game.objects.create(
            team=self.team,
            status=GameStatus.ACTIVE,
            difficulty=GameDifficulty.TOUGH,
            total_periods=10,
        )

        period = game.get_current_period_obj()

        self.assertEqual(period.P5, 20)
        self.assertEqual(period.F7, 0.6)


class DocumentApiTests(APITestCase):
    """Covers player document upload, listing, download, and validation."""

    def setUp(self):
        self.media_tmp = tempfile.TemporaryDirectory()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_tmp.name)
        self.settings_override.enable()

        session = self.client.session
        session["is_admin"] = True
        session.save()

    def tearDown(self):
        self.settings_override.disable()
        self.media_tmp.cleanup()

    def test_admin_uploads_document_and_players_download_it_as_attachment(self):
        upload = SimpleUploadedFile(
            "energy-guide.pdf",
            b"%PDF-1.4 test guide",
            content_type="application/pdf",
        )

        response = self.client.post(
            "/api/documents/",
            {
                "file": upload,
                "title": "Памятка энергетика",
                "scope": "minister",
                "minister": "energy",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertIn("download_url", response.data)
        self.assertEqual(response.data["filename"], "energy-guide.pdf")

        doc = Document.objects.get()
        list_response = self.client.get("/api/documents/")
        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            list_response.data["documents"][0]["download_url"],
            f"http://testserver/api/documents/{doc.id}/download/",
        )

        download_response = self.client.get(f"/api/documents/{doc.id}/download/")

        self.assertEqual(download_response.status_code, status.HTTP_200_OK)
        self.assertIn(
            'attachment; filename="energy-guide.pdf"',
            download_response.headers["Content-Disposition"],
        )
        self.assertEqual(
            b"".join(download_response.streaming_content),
            b"%PDF-1.4 test guide",
        )

    def test_player_documents_include_builtin_slots_when_requested(self):
        response = self.client.get("/api/documents/?include_builtin=1")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        slots = {doc["slot"]: doc for doc in response.data["documents"]}
        self.assertIn("glossary", slots)
        self.assertIn("common_mistakes", slots)
        self.assertIn("minister_population", slots)
        self.assertTrue(slots["glossary"]["is_builtin"])
        self.assertEqual(slots["glossary"]["scope"], "general")
        self.assertEqual(slots["minister_population"]["minister"], "population")
        self.assertTrue(slots["glossary"]["url"].endswith(".pdf"))

    def test_admin_can_replace_builtin_document_slot_with_uploaded_pdf(self):
        upload = SimpleUploadedFile(
            "new-glossary.pdf",
            b"%PDF-1.4 replacement",
            content_type="application/pdf",
        )

        response = self.client.post(
            "/api/documents/",
            {
                "file": upload,
                "slot": "glossary",
                "title": "Словарь терминов",
                "scope": "general",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["slot"], "glossary")
        self.assertTrue(response.data["is_builtin"])

        list_response = self.client.get("/api/documents/?include_builtin=1")
        glossary = next(
            doc for doc in list_response.data["documents"] if doc["slot"] == "glossary"
        )

        self.assertEqual(glossary["filename"], "new-glossary.pdf")
        self.assertFalse(glossary["uses_default_file"])

    def test_document_upload_accepts_only_pdf_files(self):
        upload = SimpleUploadedFile(
            "guide.docx",
            b"not a pdf",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

        response = self.client.post(
            "/api/documents/",
            {
                "file": upload,
                "title": "Памятка",
                "scope": "general",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("PDF", response.data["error"])

    def test_minister_document_upload_requires_known_minister(self):
        upload = SimpleUploadedFile(
            "guide.pdf",
            b"guide",
            content_type="application/pdf",
        )

        response = self.client.post(
            "/api/documents/",
            {"file": upload, "title": "Памятка", "scope": "minister"},
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("minister", response.data["error"])

        upload = SimpleUploadedFile(
            "guide.pdf",
            b"guide",
            content_type="application/pdf",
        )
        response = self.client.post(
            "/api/documents/",
            {
                "file": upload,
                "title": "Памятка",
                "scope": "minister",
                "minister": "unknown",
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("minister", response.data["error"])


class GameCalculatorExcelAlignmentTests(TestCase):
    """Keeps core calculations aligned with the spreadsheet reference model."""

    def setUp(self):
        self.calculator = get_calculator()
        self.calculator.reload_config()

    @staticmethod
    def _load_appendix_h_fixture() -> dict:
        fixture_path = (
            Path(__file__).resolve().parents[3]
            / "docs"
            / "regression"
            / "appendix_h_fixture.json"
        )
        with fixture_path.open(encoding="utf-8") as fixture_file:
            return json.load(fixture_file)

    def test_rounding_matches_excel_and_documented_precision(self):
        self.assertEqual(self.calculator._round_value(2.125, "G10"), 2.13)
        self.assertEqual(self.calculator._round_value(0.58, "E10"), 0.58)
        self.assertEqual(self.calculator._round_value(1810 / 5, "E9"), 360)
        self.assertEqual(self.calculator._round_value(1810 / 5, "E20"), 360)

    def test_goods_interpolation_tables_match_excel_reference_points(self):
        interpolator = self.calculator._interpolator

        self.assertEqual(interpolator.interpolate("G1", 1.0), 1.0)
        self.assertEqual(interpolator.interpolate("G3", 20.0), 14.4)

    def test_energy_interpolation_tables_match_excel_reference_curve(self):
        interpolator = self.calculator._interpolator

        self.assertEqual(interpolator.interpolate("E1", 0.0), 0.03)
        self.assertEqual(interpolator.interpolate("E1_LOW", 2.0), 20.0)

    def test_environment_recovery_edges_match_excel_percent_curve(self):
        interpolator = self.calculator._interpolator

        self.assertEqual(interpolator.interpolate("F4", 0.0), 0.0003)
        self.assertEqual(interpolator.interpolate("F4", 1.0), 0.0001)

    def test_population_formulas_use_excel_period_references(self):
        params = self.calculator.get_default_parameters()
        params.update({"P1": 220.0, "P6": 2.0, "P7": 10.0, "G6": 660.0})
        history = [{"P1": 200.0, "P7": 2.0}]

        p7 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_population"]["P7"],
            params,
            history,
        )
        p8 = self.calculator._round_value(
            self.calculator._evaluate_formula(
                self.calculator._formulas["calc_population"]["P8"],
                params,
                history,
            ),
            "P8",
        )

        self.assertEqual(p7, 3.3)
        self.assertEqual(p8, 36)

    def test_production_indicators_use_active_current_capital(self):
        params = self.calculator.get_default_parameters()
        params.update(
            {
                "F3": 2000.0,
                "F7": 0.8,
                "F10": 0.5,
                "G7": 8.0,
                "G11": 0.5,
            }
        )
        history = [{"F3": 800.0, "G3": 300.0}]

        g8 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_production_indicators"]["G8"],
            params,
            history,
        )
        f9 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_production_indicators"]["F9"],
            params,
            history,
        )

        self.assertEqual(g8, 7.2)
        self.assertEqual(f9, 5000.0)

    def test_trade_formulas_match_excel_loan_and_price_rules(self):
        params = self.calculator.get_initial_parameters(GameDifficulty.HIGHEQ)
        params["TF1"] = 100
        params["TF3"] = 1.5
        history = [
            {**params, "TF10": 100, "TF11": 200, "TF12": 700},
            {**params, "TF10": 100, "TF11": 200, "TF12": 700},
            {**params, "TF10": 100, "TF11": 200, "TF12": 700},
        ]

        max_loan = self.calculator._evaluate_formula("max_loan()", params, history)
        energy_price = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_trade_finance"]["TF6"],
            params,
            history,
        )

        self.assertEqual(max_loan, 1000)
        self.assertEqual(energy_price, 1.5)

    def test_energy_output_uses_single_excel_reference_curve(self):
        params = self.calculator.get_initial_parameters(GameDifficulty.HIGHEQ)
        params["E3"] = 2000

        e17 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_energy_balance"]["E17"],
            params,
            [params],
        )

        self.assertEqual(e17, 20000)

    def test_energy_for_population_is_not_overwritten_by_current_decision(self):
        params = self.calculator.get_default_parameters()
        params["E9"] = 400
        params["P11"] = 4500
        params["E20"] = 900

        self.calculator.apply_decision_formulas(params, [])

        self.assertEqual(params["E9"], 400)

    def test_environmental_protection_investment_does_not_go_negative(self):
        params = self.calculator.get_default_parameters()
        params.update(
            {
                "P12": 1000,
                "E26": 700,
                "E27": 400,
                "G18": 200,
                "G19": 100,
                "F14": 115,
            }
        )

        self.calculator.apply_decision_formulas(params, [])

        self.assertEqual(params["F15"], 0)

    def test_high_debt_rule_uses_energy_in_threshold_and_keeps_debt_value(self):
        params = {"TF1": 600.0, "P2": 300.0, "P3": 300.0, "E7": 1000.0}

        self.calculator._apply_special_rules(params, [])

        self.assertEqual(params["P3"], 300)
        self.assertEqual(params["TF1"], 600)

        params["TF1"] = 900.0
        self.calculator._apply_special_rules(params, [])

        self.assertEqual(params["P3"], 270)
        self.assertEqual(params["TF1"], 900)

    def test_formula_validator_accepts_configured_formula_helpers(self):
        validator = FormulaValidator()

        errors, warnings = validator.validate_all()

        self.assertEqual(errors, [])
        self.assertEqual(warnings, [])

    def test_initial_profiles_match_manual_appendix_g_core_values(self):
        expected_by_difficulty = {
            GameDifficulty.SIMPLE: {
                "P1": 200,
                "P2": 3300,
                "P3": 3500,
                "P5": 18,
                "P8": 41,
                "E7": 15000,
                "E16": 18.125,
                "E17": 14500,
                "E19": 500,
                "F7": 0.69,
                "F11": 3300,
                "G12": 3000,
                "TF1": 0,
                "TF4": 300,
                "TF5": 1000,
            },
            GameDifficulty.STANDARD: {
                "P1": 200,
                "P2": 3300,
                "P3": 3500,
                "P5": 18,
                "P8": 41,
                "E7": 14500,
                "E16": 15,
                "E17": 12000,
                "E19": 1500,
                "F7": 0.69,
                "F11": 3300,
                "G12": 3000,
                "TF1": 1000,
                "TF4": 300,
                "TF5": 0,
            },
            GameDifficulty.TOUGH: {
                "P2": 3300,
                "P5": 20,
                "E7": 14500,
                "E17": 12000,
                "F7": 0.6,
                "F11": 3300,
                "TF1": 1000,
                "TF5": 0,
            },
            GameDifficulty.VERYHARD: {
                "P2": 3000,
                "P5": 22,
                "E7": 14500,
                "E17": 12000,
                "F7": 0.5,
                "F11": 3000,
                "TF1": 1000,
                "TF5": 0,
            },
            GameDifficulty.HIGHEQ: {
                "P1": 400,
                "P2": 9000,
                "P3": 38530,
                "P4": 4.5,
                "P5": 10,
                "P6": 13.5,
                "P7": 16,
                "P8": 10,
                "E1": 1000,
                "E2": 4000,
                "E3": 5000,
                "E4": 290,
                "E5": 1160,
                "E6": 1450,
                "E7": 35400,
                "E9": 5400,
                "E16": 5,
                "E17": 25000,
                "E19": 10400,
                "F1": 320,
                "F2": 1280,
                "F3": 1600,
                "F4": 330,
                "F5": 1650,
                "F6": 1980,
                "F7": 0.8,
                "F8": 1.6,
                "F11": 8000,
                "F13": 1000,
                "G1": 400,
                "G2": 1600,
                "G3": 2000,
                "G4": 710,
                "G5": 5690,
                "G6": 6400,
                "G9": 16,
                "G12": 36530,
                "G14": 2000,
                "G16": 38530,
                "TF1": 5000,
                "TF4": 1000,
                "TF5": 5000,
            },
        }

        for difficulty, expected_values in expected_by_difficulty.items():
            with self.subTest(difficulty=difficulty):
                params = self.calculator.get_initial_parameters(difficulty)

                for param_name, expected_value in expected_values.items():
                    self.assertEqual(
                        params[param_name],
                        expected_value,
                        msg=f"{difficulty}.{param_name} differs from Appendix G",
                    )

    def test_validate_input_rejects_positive_value_when_dynamic_max_is_zero(self):
        params = self.calculator.get_default_parameters()
        params["E23"] = 0

        is_valid, error, bounds = self.calculator.validate_input(
            params,
            "E24",
            10,
            [],
        )

        self.assertFalse(is_valid)
        self.assertIsNotNone(error)
        self.assertEqual(bounds, (0.0, 0.0))

    def test_f4_and_f5_use_current_environmental_capital_like_legacy_stg(self):
        params = self.calculator.get_default_parameters()
        params["F6"] = 900.0
        history = [{"F6": 600.0}]

        f4 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_agriculture"]["F4"],
            params,
            history,
        )
        f5 = self.calculator._evaluate_formula(
            self.calculator._formulas["calc_capital"]["F5"],
            params,
            history,
        )

        self.assertEqual(f4, 150)
        self.assertEqual(f5, 750)

    def test_e27_is_limited_by_goods_and_food_investments(self):
        params = self.calculator.get_default_parameters()
        params.update(
            {
                "P12": 1000,
                "E26": 0,
                "G18": 100,
                "F14": 100,
            }
        )

        is_valid, error, bounds = self.calculator.validate_input(
            params,
            "E27",
            250,
            [],
        )

        self.assertFalse(is_valid)
        self.assertIsNotNone(error)
        self.assertEqual(bounds, (0.0, 200.0))

    def test_manual_regression_loan_limit_matches_appendix_h_history_logic(self):
        fixture = self._load_appendix_h_fixture()
        params = fixture["snapshot"]
        history = fixture["history"]

        loan_limit = self.calculator._evaluate_formula("max_loan()", params, history)

        self.assertEqual(loan_limit, 4340)

    def test_appendix_h_fixture_snapshot_matches_manual_targets(self):
        fixture = self._load_appendix_h_fixture()
        snapshot = fixture["snapshot"]
        expected = fixture["manual_expected"]

        for key, expected_value in expected.items():
            self.assertEqual(
                snapshot[key],
                expected_value,
                msg=f"{key} differs from Appendix H fixture target",
            )


class FrontendAssetTests(TestCase):
    """Guards the template-to-static-script contract used by the frontend."""

    maxDiff = None

    def setUp(self):
        self.root = Path(__file__).resolve().parents[3]
        self.frontend = self.root / "frontend"

    def read_text(self, relative_path: str) -> str:
        return (self.root / relative_path).read_text(encoding="utf-8")

    def test_page_scripts_are_extracted_from_templates(self):
        expected_assets = {
            "frontend/templates/index.html": "js/pages/team-selection.js",
            "frontend/templates/admin/login.html": "js/pages/admin-login.js",
            "frontend/templates/admin/panel.html": "js/pages/admin-panel.js",
            "frontend/templates/admin/calculator.html": "js/pages/calculator.js",
            "frontend/templates/game/docs.html": "js/pages/docs.js",
            "frontend/templates/game/play.html": "js/pages/game-play.js",
            "frontend/templates/game/minister.html": "js/pages/minister.js",
            "frontend/templates/game/charts.html": "js/pages/charts.js",
            "frontend/templates/game/status.html": "js/pages/status.js",
            "frontend/templates/game/results.html": "js/pages/results.js",
        }

        for template_path, asset_path in expected_assets.items():
            with self.subTest(template=template_path):
                template = self.read_text(template_path)
                self.assertIn(f"{{% static '{asset_path}' %}}", template)
                self.assertNotRegex(template, r"<script>\s*(?:/\*.*?\*/\s*)?function\s")
                self.assertTrue((self.frontend / "static" / asset_path).exists())

    def test_shared_api_helper_reports_specific_error_messages(self):
        app_js = self.read_text("frontend/static/js/app.js")
        base_html = self.read_text("frontend/templates/base.html")
        admin_login_template = self.read_text("frontend/templates/admin/login.html")
        admin_login_js = self.read_text("frontend/static/js/pages/admin-login.js")

        for expected_token in (
            "extractErrorMessage",
            "formatErrorPayload",
            "Ошибка запроса. Попробуйте ещё раз",
            "MAX_INLINE_ERRORS",
            "apiErrorDetails",
            "detail",
            "message",
            "non_field_errors",
            "ещё ${hiddenCount}",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, app_js)

        for expected_token in ("Показать все", "detailModal", "Ошибки в решениях"):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, base_html + app_js)

        self.assertIn("M15 12a3 3", admin_login_template)
        self.assertIn("showPassword ? 'Скрыть пароль' : 'Показать пароль'", admin_login_template)
        self.assertNotIn("admin/check", admin_login_js)

    def test_team_select_option_label_is_single_text_binding(self):
        index_template = self.read_text("frontend/templates/index.html")
        team_selection_js = self.read_text("frontend/static/js/pages/team-selection.js")

        self.assertIn('x-text="teamOptionLabel(team)"', index_template)
        self.assertNotIn('<span x-show="team.has_active_game"> (есть игра)</span>', index_template)
        self.assertIn("teamOptionLabel(team)", team_selection_js)
        self.assertIn('team.has_active_game ? `${team.name} (есть игра)` : team.name', team_selection_js)
        self.assertIn("selectedTeamHasGame", team_selection_js + index_template)
        self.assertIn('response.game.status === "finished"', team_selection_js)
        self.assertIn('"/game-results/"', team_selection_js)
        self.assertIn("Открыть итоги", index_template)
        self.assertIn('API.get("/teams/?public=1")', team_selection_js)
        self.assertIn(":type=\"showTeamPassword ? 'text' : 'password'\"", index_template)
        self.assertIn("absolute inset-y-0 right-0", index_template)
        self.assertNotIn("w-11 flex-shrink-0", index_template)
        self.assertIn("M15 12a3 3", index_template)
        self.assertIn("showTeamPassword ? 'Скрыть пароль' : 'Показать пароль'", index_template)
        self.assertIn("showTeamPassword", team_selection_js)
        self.assertIn("teamStatusMessage()", team_selection_js)
        self.assertIn("Команда завершила игру", team_selection_js)
        self.assertNotIn("Информация о команде", index_template)
        self.assertIn('autocomplete="current-password"', index_template)

    def test_admin_country_status_link_is_available_for_active_games(self):
        admin_panel = self.read_text("frontend/templates/admin/panel.html")

        country_status_link = admin_panel[
            admin_panel.index("<!-- Состояние страны -->"):
            admin_panel.index('title="Состояние страны"') + len('title="Состояние страны"')
        ]
        self.assertIn('x-show="game.status !== \'created\'"', country_status_link)
        self.assertNotIn('x-show="game.status === \'finished\'"', country_status_link)

    def test_summary_sheet_does_not_submit_locked_decisions(self):
        game_play_js = self.read_text("frontend/static/js/pages/game-play.js")
        minister_js = self.read_text("frontend/static/js/pages/minister.js")
        play_template = self.read_text("frontend/templates/game/play.html")
        minister_template = self.read_text("frontend/templates/game/minister.html")

        for expected_token in (
            "isSummaryFixed(param)",
            "is_fixed === true",
            "getParamStatus(param) === 'filled'",
            "getParamStatus(param) === 'next_to_fill'",
            "this.isSummaryInputEnabled(param)",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, game_play_js)
        self.assertIn("is_fixed === true", minister_js)
        self.assertNotIn("bounds.min === bounds.max", game_play_js)
        self.assertNotIn("bounds.min === bounds.max", minister_js)
        self.assertNotIn("b.min === b.max", game_play_js)
        self.assertIn("game-play.js' %}?v=input-locks-20260617", play_template)
        self.assertIn("minister.js' %}?v=input-locks-20260617", minister_template)

    def test_summary_sheet_renders_foreign_aid_as_summary_info(self):
        play_template = self.read_text("frontend/templates/game/play.html")
        game_play_js = self.read_text("frontend/static/js/pages/game-play.js")

        self.assertIn("SUMMARY_INFO", game_play_js)
        self.assertIn("summary-info", play_template)
        self.assertIn("Информация к сведению", play_template)
        self.assertIn("visibleSummaryInfo()", game_play_js)
        self.assertIn('x-show="visibleSummaryInfo().length > 0"', play_template)
        self.assertIn("x-for=\"param in visibleSummaryInfo()\"", play_template)
        self.assertNotIn('x-show="SUMMARY_INFO.length > 0"', play_template)
        self.assertNotIn("SUMMARY_SITUATION", game_play_js)
        self.assertNotIn("Ситуация периода", play_template)
        self.assertNotIn("group.readonly || []", play_template)

    def test_batch_http_errors_preserve_field_highlighting(self):
        app_js = self.read_text("frontend/static/js/app.js")
        game_play_js = self.read_text("frontend/static/js/pages/game-play.js")
        minister_js = self.read_text("frontend/static/js/pages/minister.js")

        expected_tokens = {
            "frontend/static/js/app.js": "error.apiErrorPayload = payload",
            "frontend/static/js/pages/game-play.js": "e.apiErrorPayload?.errors",
            "frontend/static/js/pages/minister.js": "e.apiErrorPayload?.errors",
        }
        sources = {
            "frontend/static/js/app.js": app_js,
            "frontend/static/js/pages/game-play.js": game_play_js,
            "frontend/static/js/pages/minister.js": minister_js,
        }

        for path, expected_token in expected_tokens.items():
            with self.subTest(path=path):
                self.assertIn(expected_token, sources[path])

    def test_minister_validation_banner_uses_error_color_and_plain_labels(self):
        minister_template = self.read_text("frontend/templates/game/minister.html")

        for expected_token in (
            "myErrorCount > 0 ? 'bg-red-50 border-red-200' : 'bg-yellow-50 border-yellow-200'",
            "Ошибка в решениях",
            "Не заполнено",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, minister_template)

        self.assertNotIn("Красным: ошибка в расчётах", minister_template)
        self.assertNotIn("Жёлтым: не заполнено", minister_template)

    def test_admin_panel_document_workflow_is_wired(self):
        admin_js = self.read_text("frontend/static/js/pages/admin-panel.js")
        admin_template = self.read_text("frontend/templates/admin/panel.html")
        base_template = self.read_text("frontend/templates/base.html")

        for expected_token in (
            "documents: []",
            "showUploadDocModal",
            "loadDocuments()",
            "openUploadDocModal()",
            "onDocFileChange(event)",
            "uploadDocument()",
            "deleteDocument(doc)",
            "FormData",
            "download_url",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, admin_js + admin_template)

        self.assertNotIn("стартовый файл", admin_js + admin_template)
        self.assertNotIn("Заменить стартовый файл", admin_js + admin_template)
        self.assertIn("showTeamPasswordModal", admin_js + admin_template)
        self.assertIn("openTeamPasswordModal(team)", admin_js + admin_template)
        self.assertIn("teamPasswordModalValue", admin_js + admin_template)
        self.assertIn("canSaveTeamPassword()", admin_js + admin_template)
        self.assertIn("password_enabled: this.teamPasswordModalEnabled", admin_js)
        self.assertIn("toggleTeamPasswordModalEnabled()", admin_js + admin_template)
        self.assertIn("translate-x-5", admin_template)
        self.assertIn("Оставьте поле пустым, чтобы не менять текущий пароль.", admin_template)
        self.assertNotIn("show_access_password", admin_js + admin_template)
        self.assertNotIn("x-model=\"team.access_password\"", admin_template)
        self.assertIn("showCreateTeamPassword", admin_js + admin_template)
        self.assertIn("Создать пароль", admin_template)
        self.assertIn('type="text"', admin_template)
        self.assertIn("access_password: \"\"", admin_js)
        self.assertIn("logoutAdmin", admin_template)
        self.assertIn('href="/api/docs/"', admin_template)
        self.assertNotIn("API Docs", base_template)
        self.assertIn('x-init="initialize()"', admin_template)
        self.assertIn("async initialize()", admin_js)
        self.assertNotIn("async init()", admin_js)
        self.assertIn('x-show="game.status !== \'finished\'"', admin_template)
        self.assertIn('x-show="game.status === \'finished\'"', admin_template)

        delete_entity_start = admin_js.index("async deleteEntity()")
        delete_entity_body = admin_js[delete_entity_start : admin_js.index("openUploadDocModal", delete_entity_start)]
        self.assertIn("this.loadGames()", delete_entity_body)

    def test_docs_screen_uses_role_default_and_minister_charts(self):
        docs_js = self.read_text("frontend/static/js/pages/docs.js")
        docs_template = self.read_text("frontend/templates/game/docs.html")

        for expected_token in (
            "API.get('/games/decision-structure/')",
            "this.activeDoc = this.defaultActiveDoc()",
            "if (!this.ministerKey) return null",
            "return this.ministerDocs[0] || null",
            "this.ministerConfig?.interpolation_keys",
            "getPdfJs()",
            "pdfjs.getDocument",
            "pdf.worker.min.mjs",
            "renderPdfPage(pdf, pageNumber, container, token)",
            "changePdfZoom(delta)",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, docs_js)

        for expected_token in (
            "Описание роли",
            "Описание ролей министров",
            "Словарь терминов и основные ошибки",
            "Выберите документ в списке",
            "M7 3h7l5 5v13H7z",
            "x-ref=\"pdfViewer\"",
            "height: max(720px, calc(100vh - 220px));",
            "changePdfZoom(-0.1)",
            "resetPdfZoom()",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, docs_template)

        self.assertNotIn('target="_blank"', docs_template)
        self.assertNotIn("<iframe", docs_template)
        self.assertNotIn("pdfPreviewUrl", docs_js + docs_template)
        self.assertNotIn("toolbar=0", docs_js + docs_template)
        self.assertNotIn("bg-gray-100 flex items-center justify-center", docs_template)
        self.assertNotIn("min-height: 900px", docs_template)
        self.assertNotIn("generalDocsOpen", docs_js + docs_template)
        self.assertNotIn("Общий документ", docs_template)
        self.assertNotIn("Описание роли ·", docs_template)
        self.assertNotIn("doc.filename", docs_template)

    def test_results_screen_omits_available_energy_balance(self):
        results_js = self.read_text("frontend/static/js/pages/results.js")
        results_template = self.read_text("frontend/templates/game/results.html")

        forbidden_tokens = (
            "Доступный энергобаланс",
            "Главная оценка использует доступные энергоресурсы",
            "Энергобаланс",
            "energyAvailabilityBalance",
            "energyCoverage",
            "energyBreakdown",
        )
        for forbidden_token in forbidden_tokens:
            with self.subTest(token=forbidden_token):
                self.assertNotIn(forbidden_token, results_js + results_template)

    def test_results_screen_shows_summary_score_chart_above_country_metrics(self):
        results_js = self.read_text("frontend/static/js/pages/results.js")
        results_template = self.read_text("frontend/templates/game/results.html")

        for expected_token in (
            "get scoreIndicator()",
            "scoreValue(period)",
            "renderScoreChart()",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, results_js)

        for expected_token in (
            "Сводный счёт",
            "P6 + 4·P4",
            "chart-score",
        ):
            with self.subTest(token=expected_token):
                self.assertIn(expected_token, results_template)

        score_chart_position = results_template.find("chart-score")
        key_metrics_position = results_template.find("<!-- Key Metrics Summary -->")
        self.assertGreaterEqual(score_chart_position, 0)
        self.assertGreaterEqual(key_metrics_position, 0)
        self.assertLess(score_chart_position, key_metrics_position)

    def test_templates_use_local_frontend_assets(self):
        response = self.client.get("/")
        html = response.content.decode("utf-8")

        self.assertIn("/static/css/tailwind.css", html)
        self.assertIn("/static/vendor/alpinejs/alpine.min.js", html)
        self.assertIn("/static/vendor/chartjs/chart.umd.min.js", html)
        self.assertTrue((self.frontend / "static" / "css" / "tailwind.css").exists())
        self.assertTrue(
            (self.frontend / "static" / "vendor" / "alpinejs" / "alpine.min.js").exists()
        )
        self.assertTrue(
            (self.frontend / "static" / "vendor" / "chartjs" / "chart.umd.min.js").exists()
        )

    def test_templates_do_not_reference_runtime_cdns(self):
        checked_paths = [
            *(self.frontend / "templates").rglob("*.html"),
            *(self.frontend / "static" / "js").rglob("*.js"),
        ]
        combined = "\n".join(path.read_text(encoding="utf-8") for path in checked_paths)

        for forbidden in (
            "cdn.tailwindcss.com",
            "cdn.jsdelivr.net",
            "unpkg.com",
            "cdnjs.cloudflare.com",
        ):
            with self.subTest(forbidden=forbidden):
                self.assertNotIn(forbidden, combined)

    def test_decision_ui_uses_decision_terminology(self):
        checked_paths = [
            *(self.frontend / "templates").rglob("*.html"),
            *(self.frontend / "static" / "js").rglob("*.js"),
            self.root / "backend" / "apps" / "game" / "views" / "admin.py",
        ]
        combined = "\n".join(path.read_text(encoding="utf-8") for path in checked_paths)

        old_term_plural = "пол" + "я"
        old_term_single = "пол" + "е"
        forbidden_fragments = (
            f"Все {old_term_plural} министра заполнены",
            f"{old_term_plural.capitalize()} министра заполнены",
            f"Заполните {old_term_plural} министра",
            f"Все обязательные {old_term_plural} заполнены",
            f"Заполните обязательные {old_term_plural}",
            f"Заполните все обязательные {old_term_plural}",
            f"Заполните хотя бы одно {old_term_single}",
        )
        for fragment in forbidden_fragments:
            with self.subTest(fragment=fragment):
                self.assertNotIn(fragment, combined)

        self.assertIn("Все решения заполнены", combined)
