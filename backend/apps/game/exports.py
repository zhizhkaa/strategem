"""Export helpers for game result files."""

from __future__ import annotations

from io import BytesIO
from datetime import timezone, timedelta
from typing import Any

from apps.management.models import Group

from .configuration import get_calculator_for_game
from .models import Game, GamePeriod


SUMMARY_PARAMS = [
    ("P1", "Население"),
    ("P4", "Продовольствие на человека"),
    ("P6", "Товары на человека"),
    ("F7", "Экология"),
    ("TF1", "Внешний долг"),
    ("P5", "Смертность"),
    ("P8", "Рождаемость"),
]
MSK = timezone(timedelta(hours=3))


def _periods(game: Game) -> list[GamePeriod]:
    return list(game.periods.order_by("period_number"))


def _number(value: Any) -> float | int | str:
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _score(period: GamePeriod) -> float:
    return float(period.P6 or 0) + 4 * float(period.P4 or 0)


def _format_datetime(value: Any) -> str:
    if not value:
        return ""
    return value.astimezone(MSK).strftime("%d.%m.%Y %H:%M")


def _finished_at(game: Game) -> Any:
    if game.finished_at:
        return game.finished_at
    if game.status == "finished":
        return game.updated_at
    return None


def build_export_context(game: Game) -> dict[str, Any]:
    periods = _periods(game)
    final_period = periods[-1] if periods else game.get_current_period_obj()
    score_values = [_score(period) for period in periods]
    average_score = sum(score_values) / len(score_values) if score_values else 0

    return {
        "game": game,
        "periods": periods,
        "final_period": final_period,
        "average_score": average_score,
        "summary": [
            (code, label, _number(getattr(final_period, code, "")))
            for code, label in SUMMARY_PARAMS
        ],
    }


def build_results_xlsx(game: Game) -> bytes:
    """Build a detailed XLSX workbook with all saved period parameters."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    context = build_export_context(game)
    periods: list[GamePeriod] = context["periods"]
    calculator = get_calculator_for_game(game)
    parameter_config = calculator.get_parameters_config()

    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="DDEAF7")
    subheader_fill = PatternFill("solid", fgColor="E5E7EB")
    title_fill = PatternFill("solid", fgColor="0F8FCF")
    header_font = Font(bold=True)
    title_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    metadata_rows = [
        ("Команда", game.team.name),
        ("Группа", game.team.group.name if game.team.group_id else ""),
        (
            "Факультет",
            game.team.group.faculty.name if game.team.group_id else "",
        ),
        ("Сложность", game.get_difficulty_display()),
        ("Период", f"{game.current_period} / {game.total_periods}"),
        ("Создана", _format_datetime(game.created_at)),
        ("Обновлена", _format_datetime(game.updated_at)),
        ("Закончена", _format_datetime(_finished_at(game))),
    ]

    def write_metadata(sheet, title: str) -> int:
        sheet.append((title,))
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        sheet["A1"].fill = title_fill
        sheet["A1"].font = title_font
        sheet["A1"].alignment = Alignment(horizontal="center")
        for label, value in metadata_rows:
            sheet.append((label, value))
            sheet.cell(row=sheet.max_row, column=1).font = header_font
        sheet.append(())
        return sheet.max_row + 1

    status_sheet = workbook.active
    status_sheet.title = "Состояние страны"
    status_start = write_metadata(status_sheet, "Состояние страны")
    status_headers = [
        "Показатель",
        *[f"Период {period.period_number}" for period in periods],
    ]
    status_sheet.append(status_headers)
    for cell in status_sheet[status_start]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    summary_metrics = [
        ("Сводный счёт", lambda period: _score(period)),
        ("Население", lambda period: _number(period.P1)),
        ("Продовольствие/чел", lambda period: _number(period.P4)),
        ("Товары/чел", lambda period: _number(period.P6)),
        ("Экология", lambda period: _number(period.F7)),
        ("Внешний долг", lambda period: _number(period.TF1)),
        ("Смертность", lambda period: _number(period.P5)),
        ("Рождаемость", lambda period: _number(period.P8)),
    ]
    for label, getter in summary_metrics:
        status_sheet.append([label, *[getter(period) for period in periods]])
        for cell in status_sheet[status_sheet.max_row]:
            cell.border = border
            if cell.column >= 2:
                cell.alignment = Alignment(horizontal="right")

    status_sheet.freeze_panes = f"B{status_start + 1}"
    status_sheet.column_dimensions["A"].width = 24
    for column in range(2, len(status_headers) + 1):
        status_sheet.column_dimensions[get_column_letter(column)].width = 14

    periods_sheet = workbook.create_sheet("Периоды")
    period_start = write_metadata(periods_sheet, "Подробные данные по периодам")
    period_headers = ["Код", "Показатель", *[
        f"Период {period.period_number}" for period in periods
    ]]
    periods_sheet.append(period_headers)
    for cell in periods_sheet[period_start]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for code, config in parameter_config.items():
        if not isinstance(config, dict):
            config = {"default": config}
        periods_sheet.append(
            [
                code,
                config.get("verbose_name", code),
                *[_number(getattr(period, code, "")) for period in periods],
            ]
        )
        for cell in periods_sheet[periods_sheet.max_row]:
            cell.border = border
            if cell.column >= 3:
                cell.alignment = Alignment(horizontal="right")
    periods_sheet.freeze_panes = f"C{period_start + 1}"
    periods_sheet.column_dimensions["A"].width = 10
    periods_sheet.column_dimensions["B"].width = 48
    for index in range(3, len(period_headers) + 1):
        periods_sheet.column_dimensions[get_column_letter(index)].width = 14

    for sheet in (status_sheet, periods_sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="center",
                )
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.row < 10 and cell.value and not cell.fill.fill_type:
                    cell.fill = subheader_fill

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_group_results_xlsx(group: Group) -> bytes:
    """Build one XLSX workbook with all games for a student group."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    games = list(
        Game.objects.filter(team__group=group)
        .select_related("team__group__faculty")
        .prefetch_related("periods")
        .order_by("team__name", "id")
    )

    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="DDEAF7")
    title_fill = PatternFill("solid", fgColor="0F8FCF")
    header_font = Font(bold=True)
    title_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    overview = workbook.active
    overview.title = "Игры группы"
    overview.append((f"Игры группы {group.name}",))
    overview.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    overview["A1"].fill = title_fill
    overview["A1"].font = title_font
    overview["A1"].alignment = Alignment(horizontal="center")
    overview.append(())
    headers = [
        "ID",
        "Команда",
        "Статус",
        "Сложность",
        "Период",
        "Конфигурация",
        "Сводный счёт",
        "Население",
        "Продовольствие/чел",
        "Товары/чел",
        "Внешний долг",
    ]
    overview.append(headers)
    for cell in overview[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for game in games:
        periods = _periods(game)
        final_period = periods[-1] if periods else game.get_current_period_obj()
        overview.append(
            [
                game.id,
                game.team.name,
                game.get_status_display(),
                game.get_difficulty_display(),
                f"{game.current_period} / {game.total_periods}",
                game.config_snapshot_label,
                _number(_score(final_period)),
                _number(final_period.P1),
                _number(final_period.P4),
                _number(final_period.P6),
                _number(final_period.TF1),
            ]
        )
        for cell in overview[overview.max_row]:
            cell.border = border
            if cell.column >= 7:
                cell.alignment = Alignment(horizontal="right")

    widths = [10, 28, 16, 16, 14, 20, 16, 16, 20, 16, 16]
    for index, width in enumerate(widths, start=1):
        overview.column_dimensions[get_column_letter(index)].width = width

    for game in games:
        periods = _periods(game)
        if not periods:
            continue
        sheet_title = f"{game.id}-{game.team.name}"[:31]
        sheet = workbook.create_sheet(sheet_title)
        sheet.append((f"{game.team.name} · игра {game.id}",))
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(periods))
        sheet["A1"].fill = title_fill
        sheet["A1"].font = title_font
        sheet["A1"].alignment = Alignment(horizontal="center")
        sheet.append(())
        headers = ["Код", "Название", *[f"Период {p.period_number}" for p in periods]]
        sheet.append(headers)
        for cell in sheet[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        calculator = get_calculator_for_game(game)
        for code, config in calculator.get_parameters_config().items():
            if not isinstance(config, dict):
                config = {"default": config}
            sheet.append(
                [
                    code,
                    config.get("verbose_name", code),
                    *[_number(getattr(period, code, "")) for period in periods],
                ]
            )
            for cell in sheet[sheet.max_row]:
                cell.border = border
                if cell.column >= 3:
                    cell.alignment = Alignment(horizontal="right")
        sheet.freeze_panes = "C4"
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 46
        for column in range(3, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 14

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
